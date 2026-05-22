"""
Integration tests for password import/export functionality.

Tests cover:
- CSV import with various formats
- JSON import with various formats
- XLSX export with various options
- Export with passwords included
- Export without passwords
- Export with folder and tag filtering
- Error handling and edge cases
"""

import pytest
import json
import base64
import csv
import io
from openpyxl import load_workbook

from authentication.models import PasswordEntry, PasswordFolder, PasswordTag


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordImportCSV:
    """Tests for CSV import functionality."""

    def test_import_csv_basic(self, authenticated_client, user, test_password):
        """Test importing basic CSV with passwords."""
        csv_data = """name,password,url,username,notes
Gmail,SecurePass123!,https://gmail.com,user@example.com,Personal email
GitHub,CodePass456!,https://github.com,developer,Work account
Netflix,StreamPass789!,https://netflix.com,viewer@example.com,Entertainment"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 3 passwords" in response.json()["message"]

        entries = PasswordEntry.objects.filter(user=user)
        assert entries.count() == 3

        gmail_entry = entries.get(name="Gmail")
        assert gmail_entry.site == "https://gmail.com"
        assert gmail_entry.username == "user@example.com"
        assert gmail_entry.notes == "Personal email"

    def test_import_csv_with_site_column(self, authenticated_client, user, test_password):
        """Test importing CSV with 'site' column instead of 'url'."""
        csv_data = """name,password,site,username,notes
Facebook,SocialPass123!,https://facebook.com,socialuser,Social media"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        entry = PasswordEntry.objects.get(user=user, name="Facebook")
        assert entry.site == "https://facebook.com"

    def test_import_csv_with_email_column(self, authenticated_client, user, test_password):
        """Test importing CSV with 'email' column instead of 'username'."""
        csv_data = """name,password,url,email,notes
Twitter,TweetPass123!,https://twitter.com,tweeter@example.com,Social"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        entry = PasswordEntry.objects.get(user=user, name="Twitter")
        assert entry.username == "tweeter@example.com"

    def test_import_csv_minimal_fields(self, authenticated_client, user, test_password):
        """Test importing CSV with only required fields."""
        csv_data = """name,password
Minimal,MinimalPass123!"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        entry = PasswordEntry.objects.get(user=user, name="Minimal")
        assert entry.site == ""
        assert entry.username == ""
        assert entry.notes == ""

    def test_import_csv_skip_missing_required_fields(
        self, authenticated_client, user, test_password
    ):
        """Test that rows without required fields are skipped."""
        csv_data = """name,password,url
Complete,CompletePass123!,https://example.com
Missing Password,,https://test.com
,MissingName123!,https://other.com
Valid,ValidPass456!,https://valid.com"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 2 passwords" in response.json()["message"]

        entries = PasswordEntry.objects.filter(user=user)
        assert entries.count() == 2
        assert entries.filter(name="Complete").exists()
        assert entries.filter(name="Valid").exists()

    def test_import_csv_skip_duplicates(self, authenticated_client, user, test_password):
        """Test that duplicate entries are skipped."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        encrypted_password, salt = encryption_service.encrypt_password(
            "ExistingPass123!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Existing",
            site="https://existing.com",
            encrypted_password=encrypted_password,
            encryption_salt=salt,
        )

        csv_data = """name,password,url
Existing,NewPass456!,https://existing.com
New,NewPass789!,https://new.com"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        entries = PasswordEntry.objects.filter(user=user)
        assert entries.count() == 2

    def test_import_csv_invalid_master_password(
        self, authenticated_client, user
    ):
        """Test import fails with invalid master password."""
        csv_data = """name,password
Test,TestPass123!"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": "WrongPassword123!",
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "Invalid master password" in response.json()["detail"]

    def test_import_csv_without_authentication(self, api_client):
        """Test import fails without authentication."""
        csv_data = """name,password
Test,TestPass123!"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": "Password123!",
        }

        response = api_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 401

    def test_import_csv_invalid_base64(self, authenticated_client, test_password):
        """Test import fails with invalid base64 data."""
        import_request = {
            "format": "csv",
            "data": "invalid-base64-data!!!",
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "Import failed" in response.json()["detail"]

    def test_import_csv_empty_file(self, authenticated_client, user, test_password):
        """Test importing empty CSV file."""
        csv_data = """name,password"""

        encoded_data = base64.b64encode(csv_data.encode()).decode()

        import_request = {
            "format": "csv",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 0 passwords" in response.json()["message"]


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordImportXLSX:
    """Tests for XLSX import functionality."""

    def test_import_xlsx_basic(self, authenticated_client, user, test_password):
        """Test importing basic XLSX with passwords."""
        from openpyxl import Workbook

        # Create XLSX file
        wb = Workbook()
        ws = wb.active
        ws.title = "Passwords"

        # Add headers
        ws.append(["Name", "Password", "Site", "Username", "Notes"])

        # Add data rows
        ws.append(["Gmail", "SecurePass123!", "https://gmail.com", "user@example.com", "Personal email"])
        ws.append(["GitHub", "CodePass456!", "https://github.com", "developer", "Work account"])
        ws.append(["Netflix", "StreamPass789!", "https://netflix.com", "viewer@example.com", "Entertainment"])

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 3 passwords" in response.json()["message"]

        entries = PasswordEntry.objects.filter(user=user)
        assert entries.count() == 3

        gmail_entry = entries.get(name="Gmail")
        assert gmail_entry.site == "https://gmail.com"
        assert gmail_entry.username == "user@example.com"
        assert gmail_entry.notes == "Personal email"

    def test_import_xlsx_with_url_column(self, authenticated_client, user, test_password):
        """Test importing XLSX with 'URL' column instead of 'Site'."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # Add headers with URL instead of Site
        ws.append(["Name", "Password", "URL", "Username", "Notes"])
        ws.append(["Facebook", "SocialPass123!", "https://facebook.com", "socialuser", "Social media"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        entry = PasswordEntry.objects.get(user=user, name="Facebook")
        assert entry.site == "https://facebook.com"

    def test_import_xlsx_with_email_column(self, authenticated_client, user, test_password):
        """Test importing XLSX with 'Email' column instead of 'Username'."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Password", "Site", "Email", "Notes"])
        ws.append(["Twitter", "TweetPass123!", "https://twitter.com", "tweeter@example.com", "Social"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        entry = PasswordEntry.objects.get(user=user, name="Twitter")
        assert entry.username == "tweeter@example.com"

    def test_import_xlsx_minimal_fields(self, authenticated_client, user, test_password):
        """Test importing XLSX with only required fields."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Password"])
        ws.append(["Minimal", "MinimalPass123!"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        entry = PasswordEntry.objects.get(user=user, name="Minimal")
        assert entry.site == ""
        assert entry.username == ""

    def test_import_xlsx_missing_name_column(self, authenticated_client, user, test_password):
        """Test that import fails when Name column is missing."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Password", "Site", "Username"])
        ws.append(["Pass123!", "https://example.com", "user"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "must contain" in response.json()["detail"].lower()

    def test_import_xlsx_missing_password_column(self, authenticated_client, user, test_password):
        """Test that import fails when Password column is missing."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Site", "Username"])
        ws.append(["TestEntry", "https://example.com", "user"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "must contain" in response.json()["detail"].lower()

    def test_import_xlsx_skip_empty_rows(self, authenticated_client, user, test_password):
        """Test that empty rows are skipped during import."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Password", "Site"])
        ws.append(["Valid", "Pass123!", "https://example.com"])
        ws.append(["", "", ""])  # Empty row
        ws.append([None, None, None])  # Null row
        ws.append(["Another", "Pass456!", "https://test.com"])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 2 passwords" in response.json()["message"]

    def test_import_xlsx_skip_duplicates(self, authenticated_client, user, test_password):
        """Test that duplicate entries are skipped."""
        from authentication.encryption_service import encryption_service

        # Create existing entry
        encrypted_password, salt = encryption_service.encrypt_password(
            "ExistingPass!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Existing",
            site="https://existing.com",
            encrypted_password=encrypted_password,
            encryption_salt=salt,
        )

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Password", "Site"])
        ws.append(["Existing", "NewPass123!", "https://existing.com"])  # Duplicate
        ws.append(["New", "NewPass456!", "https://new.com"])  # New entry

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        # Verify only the new entry was added
        new_entries = PasswordEntry.objects.filter(user=user)
        assert new_entries.count() == 2

    def test_import_xlsx_with_none_values(self, authenticated_client, user, test_password):
        """Test importing XLSX with None/null values in optional fields."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        ws.append(["Name", "Password", "Site", "Username", "Notes"])
        ws.append(["TestEntry", "Pass123!", None, None, None])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        encoded_data = base64.b64encode(output.getvalue()).decode()

        import_request = {
            "format": "xlsx",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

        entry = PasswordEntry.objects.get(user=user, name="TestEntry")
        assert entry.site == ""
        assert entry.username == ""
        assert entry.notes == ""


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordImportJSON:
    """Tests for JSON import functionality."""

    def test_import_json_basic(self, authenticated_client, user, test_password):
        """Test importing basic JSON with passwords."""
        json_data = [
            {
                "name": "Gmail",
                "password": "SecurePass123!",
                "site": "https://gmail.com",
                "username": "user@example.com",
                "notes": "Personal email",
            },
            {
                "name": "GitHub",
                "password": "CodePass456!",
                "site": "https://github.com",
                "username": "developer",
                "notes": "Work account",
            },
        ]

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 2 passwords" in response.json()["message"]

        entries = PasswordEntry.objects.filter(user=user)
        assert entries.count() == 2

    def test_import_json_single_object(self, authenticated_client, user, test_password):
        """Test importing single JSON object (not array)."""
        json_data = {
            "name": "Solo",
            "password": "SoloPass123!",
            "site": "https://solo.com",
            "username": "solo_user",
        }

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

    def test_import_json_minimal_fields(self, authenticated_client, user, test_password):
        """Test importing JSON with only required fields."""
        json_data = [{"name": "Minimal", "password": "MinimalPass123!"}]

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        entry = PasswordEntry.objects.get(user=user, name="Minimal")
        assert entry.site == ""
        assert entry.username == ""

    def test_import_json_skip_missing_required_fields(
        self, authenticated_client, user, test_password
    ):
        """Test that entries without required fields are skipped."""
        json_data = [
            {"name": "Valid", "password": "ValidPass123!", "site": "https://valid.com"},
            {"password": "NoName123!", "site": "https://test.com"},
            {"name": "NoPassword", "site": "https://test2.com"},
            {
                "name": "Another",
                "password": "AnotherPass456!",
                "site": "https://another.com",
            },
        ]

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 2 passwords" in response.json()["message"]

    def test_import_json_skip_duplicates(self, authenticated_client, user, test_password):
        """Test that duplicate entries are skipped."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        encrypted_password, salt = encryption_service.encrypt_password(
            "ExistingPass123!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Existing",
            site="https://existing.com",
            encrypted_password=encrypted_password,
            encryption_salt=salt,
        )

        json_data = [
            {
                "name": "Existing",
                "password": "NewPass456!",
                "site": "https://existing.com",
            },
            {"name": "New", "password": "NewPass789!", "site": "https://new.com"},
        ]

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 1 passwords" in response.json()["message"]

    def test_import_json_invalid_json(self, authenticated_client, test_password):
        """Test import fails with invalid JSON."""
        invalid_json = "{ this is not valid json }"
        encoded_data = base64.b64encode(invalid_json.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "Import failed" in response.json()["detail"]

    def test_import_json_empty_array(self, authenticated_client, user, test_password):
        """Test importing empty JSON array."""
        json_data = []

        json_string = json.dumps(json_data)
        encoded_data = base64.b64encode(json_string.encode()).decode()

        import_request = {
            "format": "json",
            "data": encoded_data,
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        assert "Successfully imported 0 passwords" in response.json()["message"]


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordExportWithPasswords:
    """Tests for password export with passwords included."""

    def test_export_json_with_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in JSON format with passwords included."""
        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        data = response.json()

        assert data["format"] == "json"
        assert data["filename"] == "passwords_export.json"
        assert "data" in data

        decoded_data = base64.b64decode(data["data"]).decode()
        export_data = json.loads(decoded_data)

        assert isinstance(export_data, list)
        assert len(export_data) >= 1

        exported_entry = next(
            (e for e in export_data if e["name"] == password_entry.name), None
        )
        assert exported_entry is not None
        assert "password" in exported_entry
        assert exported_entry["site"] == password_entry.site
        assert exported_entry["username"] == password_entry.username

    def test_export_csv_with_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in CSV format with passwords included."""
        export_request = {
            "format": "csv",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        data = response.json()

        assert data["format"] == "csv"
        assert data["filename"] == "passwords_export.csv"

        decoded_data = base64.b64decode(data["data"]).decode()
        csv_reader = csv.DictReader(io.StringIO(decoded_data))
        rows = list(csv_reader)

        assert len(rows) >= 1
        assert "password" in rows[0]
        assert "name" in rows[0]
        assert "site" in rows[0]
        assert "username" in rows[0]

    def test_export_xlsx_with_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in XLSX format with passwords included."""
        export_request = {
            "format": "xlsx",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200
        data = response.json()

        assert data["format"] == "xlsx"
        assert data["filename"] == "passwords_export.xlsx"
        assert "data" in data

        # Decode and load the XLSX file
        decoded_data = base64.b64decode(data["data"])
        wb = load_workbook(io.BytesIO(decoded_data))
        ws = wb.active

        # Verify worksheet name
        assert ws.title == "Passwords"

        # Verify headers exist (row 1)
        headers = [cell.value for cell in ws[1]]
        assert "Name" in headers
        assert "Site" in headers
        assert "Username" in headers
        assert "Notes" in headers
        assert "Password" in headers

        # Verify at least one data row exists (row 2+)
        assert ws.max_row >= 2

        # Verify header styling
        header_cell = ws.cell(row=1, column=1)
        # Color can be stored with or without alpha channel prefix
        assert header_cell.fill.start_color.rgb in ["FF4F46E5", "004F46E5"]  # Indigo background
        assert header_cell.font.bold is True
        assert header_cell.font.color.rgb in ["FFFFFFFF", "00FFFFFF"]  # White text

    def test_export_xlsx_without_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in XLSX format without passwords."""
        export_request = {
            "format": "xlsx",
            "master_password": test_password,
            "include_passwords": False,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        # Decode and load the XLSX file
        decoded_data = base64.b64decode(response.json()["data"])
        wb = load_workbook(io.BytesIO(decoded_data))
        ws = wb.active

        # Verify headers
        headers = [cell.value for cell in ws[1]]
        assert "Name" in headers
        assert "Site" in headers
        assert "Username" in headers
        assert "Notes" in headers
        assert "Password" not in headers  # Password column should not be present

    def test_export_xlsx_multiple_entries(
        self, authenticated_client, user, test_password
    ):
        """Test exporting multiple entries to XLSX format."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        # Create test entries
        test_data = [
            {"name": "Gmail", "site": "https://gmail.com", "username": "user1@gmail.com", "password": "Pass1!"},
            {"name": "GitHub", "site": "https://github.com", "username": "dev_user", "password": "Pass2!"},
            {"name": "LinkedIn", "site": "https://linkedin.com", "username": "professional", "password": "Pass3!"},
        ]

        for entry_data in test_data:
            encrypted_password, salt = encryption_service.encrypt_password(
                entry_data["password"], test_password
            )
            PasswordEntry.objects.create(
                user=user,
                name=entry_data["name"],
                site=entry_data["site"],
                username=entry_data["username"],
                encrypted_password=encrypted_password,
                encryption_salt=salt,
                notes=f"Notes for {entry_data['name']}",
            )

        export_request = {
            "format": "xlsx",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        # Decode and load the XLSX file
        decoded_data = base64.b64decode(response.json()["data"])
        wb = load_workbook(io.BytesIO(decoded_data))
        ws = wb.active

        # Verify row count (1 header + 3 data rows)
        assert ws.max_row >= 4

        # Verify data is present
        row_2_name = ws.cell(row=2, column=1).value
        assert row_2_name in [entry["name"] for entry in test_data]

    def test_export_xlsx_empty_entries(
        self, authenticated_client, user, test_password
    ):
        """Test exporting to XLSX when no entries exist."""
        export_request = {
            "format": "xlsx",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        # Decode and load the XLSX file
        decoded_data = base64.b64decode(response.json()["data"])
        wb = load_workbook(io.BytesIO(decoded_data))
        ws = wb.active

        # Should have headers but no data rows
        assert ws.max_row == 1  # Only header row

    def test_export_xlsx_with_column_width_exception(
        self, authenticated_client, user, test_password
    ):
        """Test XLSX export handles exceptions during column width calculation (lines 728-729)."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service
        from unittest.mock import patch, PropertyMock

        # Create entry with data
        encrypted_password, salt = encryption_service.encrypt_password(
            "TestPass123!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Test Entry",
            site="https://example.com",
            username="testuser",
            encrypted_password=encrypted_password,
            encryption_salt=salt,
        )

        export_request = {
            "format": "xlsx",
            "master_password": test_password,
            "include_passwords": True,
        }

        # Create a mock cell that raises exception when value property is accessed
        # This will trigger the exception handler in lines 728-729
        class MockCell:
            @property
            def value(self):
                raise RuntimeError("Simulated cell value access error")

        original_iter_rows = None
        call_count = [0]

        def mock_iter_rows(self, min_col=None, max_col=None, min_row=None, max_row=None):
            # Only inject the problematic cell on subsequent calls (during width calculation)
            call_count[0] += 1
            if call_count[0] > 1 and min_row == 2:
                # Return a problematic cell to trigger the exception
                return iter([[MockCell()]])
            # For other calls, use the original method
            if original_iter_rows:
                return original_iter_rows(self, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
            return iter([])

        with patch('openpyxl.worksheet.worksheet.Worksheet.iter_rows', mock_iter_rows):
            response = authenticated_client.post(
                "/api/passwords/export",
                data=json.dumps(export_request),
                content_type="application/json",
            )

            # Should still succeed despite exception in width calculation
            assert response.status_code == 200
            assert response.json()["format"] == "xlsx"

    def test_export_multiple_entries_with_passwords(
        self, authenticated_client, user, test_password
    ):
        """Test exporting multiple password entries."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        for i in range(3):
            encrypted_password, salt = encryption_service.encrypt_password(
                f"TestPass{i}123!", test_password
            )
            PasswordEntry.objects.create(
                user=user,
                name=f"Entry {i}",
                site=f"https://site{i}.com",
                username=f"user{i}",
                encrypted_password=encrypted_password,
                encryption_salt=salt,
            )

        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        export_data = json.loads(decoded_data)

        assert len(export_data) == 3


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordExportWithoutPasswords:
    """Tests for password export without passwords."""

    def test_export_json_without_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in JSON format without passwords."""
        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": False,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        export_data = json.loads(decoded_data)

        assert len(export_data) >= 1
        exported_entry = export_data[0]
        assert "password" not in exported_entry
        assert "name" in exported_entry
        assert "site" in exported_entry
        assert "username" in exported_entry
        assert "notes" in exported_entry

    def test_export_csv_without_passwords(
        self, authenticated_client, user, test_password, password_entry
    ):
        """Test exporting passwords in CSV format without passwords."""
        export_request = {
            "format": "csv",
            "master_password": test_password,
            "include_passwords": False,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        csv_reader = csv.DictReader(io.StringIO(decoded_data))
        rows = list(csv_reader)

        assert len(rows) >= 1
        assert "password" not in rows[0]
        assert "name" in rows[0]


@pytest.mark.integration
@pytest.mark.django_db
class TestPasswordExportFiltering:
    """Tests for export with folder and tag filtering."""

    def test_export_by_folder(self, authenticated_client, user, test_password):
        """Test exporting passwords filtered by folder."""
        from authentication.models import PasswordEntry, PasswordFolder
        from authentication.encryption_service import encryption_service

        folder = PasswordFolder.objects.create(user=user, name="Work")

        encrypted_password1, salt1 = encryption_service.encrypt_password(
            "WorkPass1!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Work Entry",
            folder=folder,
            encrypted_password=encrypted_password1,
            encryption_salt=salt1,
        )

        encrypted_password2, salt2 = encryption_service.encrypt_password(
            "PersonalPass1!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="Personal Entry",
            encrypted_password=encrypted_password2,
            encryption_salt=salt2,
        )

        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": False,
            "folder_id": str(folder.id),
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        export_data = json.loads(decoded_data)

        assert len(export_data) == 1
        assert export_data[0]["name"] == "Work Entry"

    def test_export_by_tags(self, authenticated_client, user, test_password):
        """Test exporting passwords filtered by tags."""
        from authentication.models import PasswordEntry, PasswordTag
        from authentication.encryption_service import encryption_service

        work_tag = PasswordTag.objects.create(user=user, name="work")
        personal_tag = PasswordTag.objects.create(user=user, name="personal")

        encrypted_password1, salt1 = encryption_service.encrypt_password(
            "WorkPass1!", test_password
        )
        work_entry = PasswordEntry.objects.create(
            user=user,
            name="Work Entry",
            encrypted_password=encrypted_password1,
            encryption_salt=salt1,
        )
        work_entry.tags.add(work_tag)

        encrypted_password2, salt2 = encryption_service.encrypt_password(
            "PersonalPass1!", test_password
        )
        personal_entry = PasswordEntry.objects.create(
            user=user,
            name="Personal Entry",
            encrypted_password=encrypted_password2,
            encryption_salt=salt2,
        )
        personal_entry.tags.add(personal_tag)

        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": False,
            "tag_ids": [str(work_tag.id)],
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        export_data = json.loads(decoded_data)

        assert len(export_data) == 1
        assert export_data[0]["name"] == "Work Entry"

    def test_export_empty_result(self, authenticated_client, user, test_password):
        """Test exporting when no entries match filters."""
        from authentication.models import PasswordFolder

        folder = PasswordFolder.objects.create(user=user, name="Empty")

        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": False,
            "folder_id": str(folder.id),
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 200

        decoded_data = base64.b64decode(response.json()["data"]).decode()
        export_data = json.loads(decoded_data)

        assert len(export_data) == 0


@pytest.mark.integration
@pytest.mark.django_db
class TestImportExportErrorHandling:
    """Tests for error handling in import/export."""

    def test_export_invalid_master_password(self, authenticated_client, password_entry):
        """Test export fails with invalid master password."""
        export_request = {
            "format": "json",
            "master_password": "WrongPassword123!",
            "include_passwords": True,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 400
        assert "Invalid master password" in response.json()["detail"]

    def test_export_without_authentication(self, api_client):
        """Test export fails without authentication."""
        export_request = {
            "format": "json",
            "master_password": "Password123!",
            "include_passwords": True,
        }

        response = api_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 401

    def test_import_unsupported_format(self, authenticated_client, test_password):
        """Test import fails with unsupported format."""
        import_request = {
            "format": "xml",
            "data": base64.b64encode(b"<xml></xml>").decode(),
            "master_password": test_password,
        }

        response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert response.status_code == 422

    def test_export_unsupported_format(self, authenticated_client, test_password):
        """Test export fails with unsupported format."""
        export_request = {
            "format": "xml",
            "master_password": test_password,
            "include_passwords": False,
        }

        response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert response.status_code == 422


@pytest.mark.integration
@pytest.mark.django_db
class TestImportExportRoundTrip:
    """Tests for complete import/export cycles."""

    def test_export_import_json_roundtrip(
        self, authenticated_client, user, test_password
    ):
        """Test exporting and re-importing data maintains integrity."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        original_entries = []
        for i in range(3):
            encrypted_password, salt = encryption_service.encrypt_password(
                f"TestPass{i}123!", test_password
            )
            entry = PasswordEntry.objects.create(
                user=user,
                name=f"Entry {i}",
                site=f"https://site{i}.com",
                username=f"user{i}@example.com",
                notes=f"Notes for entry {i}",
                encrypted_password=encrypted_password,
                encryption_salt=salt,
            )
            original_entries.append(entry)

        export_request = {
            "format": "json",
            "master_password": test_password,
            "include_passwords": True,
        }

        export_response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert export_response.status_code == 200

        PasswordEntry.objects.filter(user=user).delete()

        import_request = {
            "format": "json",
            "data": export_response.json()["data"],
            "master_password": test_password,
        }

        import_response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert import_response.status_code == 200
        assert "Successfully imported 3 passwords" in import_response.json()["message"]

        imported_entries = PasswordEntry.objects.filter(user=user)
        assert imported_entries.count() == 3

        for original in original_entries:
            imported = imported_entries.get(
                name=original.name, site=original.site
            )
            assert imported.username == original.username
            assert imported.notes == original.notes

    def test_export_import_csv_roundtrip(
        self, authenticated_client, user, test_password
    ):
        """Test CSV export/import roundtrip."""
        from authentication.models import PasswordEntry
        from authentication.encryption_service import encryption_service

        encrypted_password, salt = encryption_service.encrypt_password(
            "TestPass123!", test_password
        )
        PasswordEntry.objects.create(
            user=user,
            name="CSV Test",
            site="https://csvtest.com",
            username="csvuser",
            notes="CSV notes",
            encrypted_password=encrypted_password,
            encryption_salt=salt,
        )

        export_request = {
            "format": "csv",
            "master_password": test_password,
            "include_passwords": True,
        }

        export_response = authenticated_client.post(
            "/api/passwords/export",
            data=json.dumps(export_request),
            content_type="application/json",
        )

        assert export_response.status_code == 200

        PasswordEntry.objects.filter(user=user).delete()

        import_request = {
            "format": "csv",
            "data": export_response.json()["data"],
            "master_password": test_password,
        }

        import_response = authenticated_client.post(
            "/api/passwords/import",
            data=json.dumps(import_request),
            content_type="application/json",
        )

        assert import_response.status_code == 200
        assert "Successfully imported 1 passwords" in import_response.json()["message"]

        imported_entry = PasswordEntry.objects.get(user=user, name="CSV Test")
        assert imported_entry.site == "https://csvtest.com"
        assert imported_entry.username == "csvuser"
        assert imported_entry.notes == "CSV notes"
